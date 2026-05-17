import io
import csv
from datetime import date, datetime, timedelta
from calendar import monthrange

from flask import Blueprint, render_template, jsonify, request, Response
from database import db, User, AttendanceLog

attendance_bp = Blueprint('attendance', __name__)


@attendance_bp.route('/attendance')
def attendance():
    return render_template('attendance.html')


@attendance_bp.route('/api/attendance')
def api_attendance():
    view  = request.args.get('view',  'daily')     # daily | monthly | yearly
    ref   = request.args.get('date',  date.today().isoformat())
    dept  = request.args.get('dept',  '')
    search = request.args.get('q',    '').strip().lower()

    try:
        ref_date = date.fromisoformat(ref)
    except ValueError:
        ref_date = date.today()

    query = (db.session.query(AttendanceLog)
             .join(User, AttendanceLog.user_id == User.id)
             .filter(User.is_active == True))

    if dept:
        query = query.filter(User.department == dept)

    # ── Date filter ──────────────────────────────────────────────────
    if view == 'daily':
        query = query.filter(AttendanceLog.date == ref_date)
    elif view == 'monthly':
        query = query.filter(
            db.extract('year',  AttendanceLog.date) == ref_date.year,
            db.extract('month', AttendanceLog.date) == ref_date.month)
    elif view == 'yearly':
        query = query.filter(
            db.extract('year',  AttendanceLog.date) == ref_date.year)

    logs = query.order_by(AttendanceLog.date.desc(),
                          AttendanceLog.check_in.desc()).all()

    # Optional name search (post-filter)
    if search:
        logs = [l for l in logs if search in l.user.name.lower()]

    records = [l.to_dict() for l in logs]

    # ── Chart data ───────────────────────────────────────────────────
    chart = _build_chart(view, ref_date, dept)

    # ── Departments list ─────────────────────────────────────────────
    depts = [r[0] for r in db.session.query(User.department)
             .filter(User.is_active == True)
             .distinct().order_by(User.department).all() if r[0]]

    return jsonify({'records': records, 'chart': chart, 'departments': depts})


def _build_chart(view, ref_date, dept):
    """Return {labels, present, absent} for the chart."""
    q_base = (db.session.query(AttendanceLog.date,
                               db.func.count(AttendanceLog.id))
              .join(User)
              .filter(User.is_active == True))
    if dept:
        q_base = q_base.filter(User.department == dept)

    if view == 'daily':
        # Hourly counts for ref_date
        from collections import defaultdict
        hour_map = defaultdict(int)
        day_logs = (db.session.query(AttendanceLog)
                    .join(User).filter(User.is_active == True)
                    .filter(AttendanceLog.date == ref_date)
                    .filter(AttendanceLog.check_in != None).all())
        if dept:
            day_logs = [l for l in day_logs if l.user.department == dept]
        for log in day_logs:
            try:
                h = int(log.check_in.split(':')[0])
                hour_map[h] += 1
            except Exception:
                pass
        labels  = [f"{h:02d}:00" for h in range(24)]
        present = [hour_map.get(h, 0) for h in range(24)]
        return {'labels': labels, 'present': present}

    elif view == 'monthly':
        days_in_month = monthrange(ref_date.year, ref_date.month)[1]
        labels, present = [], []
        for d in range(1, days_in_month + 1):
            day = date(ref_date.year, ref_date.month, d)
            cnt = (db.session.query(db.func.count(AttendanceLog.id))
                   .join(User).filter(User.is_active == True)
                   .filter(AttendanceLog.date == day)
                   .filter(AttendanceLog.check_in != None)
                   .scalar()) or 0
            labels.append(str(d))
            present.append(cnt)
        return {'labels': labels, 'present': present}

    else:  # yearly
        labels, present = [], []
        month_names = ['Jan','Feb','Mar','Apr','May','Jun',
                       'Jul','Aug','Sep','Oct','Nov','Dec']
        for m in range(1, 13):
            cnt = (db.session.query(db.func.count(AttendanceLog.id))
                   .join(User).filter(User.is_active == True)
                   .filter(db.extract('year',  AttendanceLog.date) == ref_date.year)
                   .filter(db.extract('month', AttendanceLog.date) == m)
                   .filter(AttendanceLog.check_in != None)
                   .scalar()) or 0
            labels.append(month_names[m - 1])
            present.append(cnt)
        return {'labels': labels, 'present': present}


@attendance_bp.route('/api/attendance/export')
def export_excel():
    view     = request.args.get('view', 'daily')
    ref      = request.args.get('date', date.today().isoformat())
    try:
        ref_date = date.fromisoformat(ref)
    except ValueError:
        ref_date = date.today()

    query = (db.session.query(AttendanceLog)
             .join(User).filter(User.is_active == True))

    if view == 'daily':
        query = query.filter(AttendanceLog.date == ref_date)
    elif view == 'monthly':
        query = query.filter(
            db.extract('year',  AttendanceLog.date) == ref_date.year,
            db.extract('month', AttendanceLog.date) == ref_date.month)
    elif view == 'yearly':
        query = query.filter(
            db.extract('year',  AttendanceLog.date) == ref_date.year)

    logs = query.order_by(AttendanceLog.date, AttendanceLog.check_in).all()

    # ── Excel Generation using openpyxl ──────────────────────────────
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance ({view.title()})"

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5") # Indigo primary

    # Headers
    headers = ['Date', 'Employee ID', 'Name', 'Department', 'Role', 
               'Check-In', 'Check-Out', 'Hours Worked', 'Status']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Data Rows
    present_count, absent_count, late_count = 0, 0, 0
    
    for log in logs:
        ws.append([
            str(log.date), log.user.employee_id, log.user.name,
            log.user.department, log.user.role,
            log.check_in or '', log.check_out or '',
            log.hours_worked or '', log.status
        ])
        if log.status == 'Present': present_count += 1
        elif log.status == 'Late': late_count += 1
        elif log.status == 'Absent': absent_count += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # ── Summary Sheet ────────────────────────────────────────────────
    ws_summary = wb.create_sheet(title="Summary Report")
    ws_summary.append(["Metric", "Count"])
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    ws_summary.append(["Total Records", len(logs)])
    ws_summary.append(["Present", present_count])
    ws_summary.append(["Late Arrivals", late_count])
    ws_summary.append(["Absent", absent_count])
    
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15

    # Output to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition':
                 f'attachment; filename=FaceID_Attendance_{view}_{ref}.xlsx'}
    )
