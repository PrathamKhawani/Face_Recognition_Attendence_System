"""
Bulk import route: POST /users/import
Accepts a CSV or Excel file with columns:
  Name, Employee_ID, Department, Role, Email, Phone, Shift_Start, Shift_End, Join_Date
"""

import csv
import io
from datetime import date

from flask import Blueprint, request, jsonify, render_template
from database import db, User, AttendanceLog
import config

import_bp = Blueprint('import_users', __name__)


REQUIRED_COLS = {'name'}
OPTIONAL_COLS = {
    'employee_id', 'department', 'role',
    'email', 'phone', 'shift_start', 'shift_end', 'join_date'
}


@import_bp.route('/users/import', methods=['GET'])
def import_page():
    return render_template('import_users.html')


@import_bp.route('/users/import', methods=['POST'])
def do_import():
    file = request.files.get('csv_file')
    if not file or not file.filename:
        return jsonify({'error': 'No file uploaded'}), 400

    fname = file.filename.lower()
    if not (fname.endswith('.csv') or fname.endswith('.xlsx') or fname.endswith('.xls')):
        return jsonify({'error': 'Only CSV or Excel files are supported'}), 400

    rows = []

    # ── Parse CSV ──────────────────────────────────────────────────
    if fname.endswith('.csv'):
        try:
            content = file.read().decode('utf-8-sig')
            reader  = csv.DictReader(io.StringIO(content))
            for row in reader:
                rows.append({k.strip().lower(): v.strip() for k, v in row.items()})
        except Exception as e:
            return jsonify({'error': f'CSV parse error: {e}'}), 400

    # ── Parse Excel ────────────────────────────────────────────────
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [str(c.value).strip().lower() for c in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (str(v).strip() if v is not None else '')
                             for i, v in enumerate(row)})
        except ImportError:
            return jsonify({'error': 'openpyxl not installed — use CSV format instead'}), 500
        except Exception as e:
            return jsonify({'error': f'Excel parse error: {e}'}), 400

    if not rows:
        return jsonify({'error': 'File is empty or has no data rows'}), 400

    # ── Normalise column names ─────────────────────────────────────
    # Accept alternate spellings
    col_map = {
        'full name': 'name', 'fullname': 'name',
        'employee id': 'employee_id', 'emp id': 'employee_id', 'empid': 'employee_id',
        'dept': 'department', 'position': 'role', 'designation': 'role',
        'mobile': 'phone', 'contact': 'phone',
        'joining date': 'join_date', 'date of joining': 'join_date',
    }
    normalised = []
    for row in rows:
        n = {}
        for k, v in row.items():
            n[col_map.get(k, k)] = v
        normalised.append(n)

    imported, skipped, errors = 0, 0, []

    for i, row in enumerate(normalised, start=2):
        name = row.get('name', '').strip()
        if not name:
            skipped += 1
            continue

        emp_id = row.get('employee_id', '').strip()
        if not emp_id:
            prefix = row.get('department', 'EMP')[:3].upper()
            import uuid
            emp_id = f"{prefix}-{str(uuid.uuid4())[:6].upper()}"

        if User.query.filter_by(employee_id=emp_id).first():
            errors.append(f"Row {i}: Employee ID '{emp_id}' already exists — skipped")
            skipped += 1
            continue

        join_date = date.today()
        jd_str = row.get('join_date', '')
        if jd_str:
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
                try:
                    from datetime import datetime as dt
                    join_date = dt.strptime(jd_str, fmt).date()
                    break
                except ValueError:
                    pass

        user = User(
            name        = name,
            employee_id = emp_id,
            department  = row.get('department', ''),
            role        = row.get('role', ''),
            email       = row.get('email', ''),
            phone       = row.get('phone', ''),
            shift_start = row.get('shift_start', config.DEFAULT_SHIFT_START) or config.DEFAULT_SHIFT_START,
            shift_end   = row.get('shift_end',   config.DEFAULT_SHIFT_END)   or config.DEFAULT_SHIFT_END,
            join_date   = join_date,
            image_path  = '',
        )
        db.session.add(user)
        imported += 1

    db.session.commit()

    # Reload face engine (no images yet, but counts are updated)
    from face_engine import engine
    active = User.query.filter_by(is_active=True).all()
    engine.load_faces(active)

    return jsonify({
        'status':   'ok',
        'imported': imported,
        'skipped':  skipped,
        'errors':   errors[:20],   # cap error list
    })
